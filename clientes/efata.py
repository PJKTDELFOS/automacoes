import sys
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import re
import os
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
from engine_busca_pncp.base_monitor import BaseMonitor



class ClientesEfataEventos(BaseMonitor):
    def filtros(self,dados_brutos):
        self.dados_filtrados=[]
        self.ids_a_registrar=[]

        for item in dados_brutos:
            if self.verificar_duplicidade(item):
                continue
            objeto=item.get('objetoCompra','') or ""
            objeto_limpo=re.sub(r'[\x00-\x1F\x7F]', '', objeto)
            if any(palavra in objeto_limpo.lower() for palavra in self.palavras_chave):
                data_fim_str=item.get('dataEncerramentoProposta','')
                if data_fim_str:
                    data_fim=datetime.fromisoformat(data_fim_str.replace('Z','').split('.')[0])
                    if data_fim>=self.hoje:
                        unidade = item.get('unidadeOrgao', {}) or {}
                        registro = {
                            'DATA': data_fim.strftime('%d/%m/%Y %H:%M'),  # [cite: 507]
                            'NUMERO': f"{item.get('numeroCompra')}/{item.get('anoCompra')}",  # [cite: 504]
                            'MODALIDADE': item.get('modalidadeNome', ''),  # [cite: 507]
                            'ORGAO': item.get('orgaoEntidade', {}).get('razaoSocial', ''),  # [cite: 509]
                            'OBJETO': objeto_limpo.upper() + '' + '',  # [cite: 507]
                            'UASG': unidade.get('codigoUnidade', "")  # [cite: 509]
                        }
                        self.dados_filtrados.append(registro)
                        self.ids_a_registrar.append(self.gerar_id_unico(item))


    def gerar_planilha(self):
        if not super().gerar_planilha():
            return False
        try:
            df = pd.DataFrame(self.dados_filtrados.copy())
            if not df.empty:
                df['DATA_OBJ'] = pd.to_datetime(df['DATA'], format='%d/%m/%Y %H:%M')
                df = df.sort_values(by='DATA_OBJ', ascending=True)
                df = df.drop(columns=['DATA_OBJ'])
                df = df[['DATA', 'NUMERO', 'MODALIDADE', 'ORGAO', 'OBJETO', 'UASG']]
                df.to_excel(self.nome_arquivo, index=False)
            wb = load_workbook(self.nome_arquivo)
            ws = wb.active
            cores_modalidades = {
                'TOMADA DE PREÇOS': PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
                # Vermelho
                'PREGÃO - ELETRÔNICO': PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid"),
                # Verde
                'PREGÃO - PRESENCIAL': PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
                # Amarelo
                'CONCORRÊNCIA': PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid"),  # Laranja
                'CREDENCIAMENTO': PatternFill(start_color="800080", end_color="800080", fill_type="solid")  # Roxo
            }
            fonte_padrao = Font(name='Century Gothic', size=8)
            alinhamento = Alignment(wrap_text=True, vertical='center', horizontal='left')

            preenchimento_padrao = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                modalidade_celula = row[2]
                texto_modalidade = str(modalidade_celula.value).upper() if modalidade_celula.value else ""
                cor_localizada = False
                for cell in row:
                    cell.font = fonte_padrao
                    cell.alignment = alinhamento
                for chave, preenchimento in cores_modalidades.items():
                    if chave in texto_modalidade:
                        modalidade_celula.fill = preenchimento
                        cor_localizada = True
                        if chave in ['CREDENCIAMENTO', 'TOMADA DE PREÇOS']:
                            modalidade_celula.font = Font(name='Century Gothic', size=8, color="FFFFFF")
                        break
                if not cor_localizada:
                    modalidade_celula.fill = preenchimento_padrao

            ws.column_dimensions['A'].width = 18  # Data
            ws.column_dimensions['B'].width = 15  # Número
            ws.column_dimensions['C'].width = 25  # Modalidade
            ws.column_dimensions['D'].width = 40  # Órgão
            ws.column_dimensions['E'].width = 60  # Objeto
            ws.column_dimensions['F'].width = 12  # UASG
            wb.save(self.nome_arquivo)

            return True


        except Exception as e:
            print(f"Erro ao formatar planilha {self.cliente}: {e}")
            return False




