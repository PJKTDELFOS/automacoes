import sys
import pandas as pd
from datetime import datetime, timedelta,timezone
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import re
import os
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
from engine_busca_pncp.base_monitor import BaseMonitor



class MonitorClientes(BaseMonitor):
    def filtros(self,dados_brutos):
        self.dados_filtrados=[]
        lista_temp=[]
        for item in dados_brutos:
            objeto=item.get('objetoCompra','') or ""
            objeto_limpo = re.sub(r'[\x00-\x1F\x7F]', '', objeto).strip().upper()
            data_fim_str=item.get('dataEncerramentoProposta','')
            agora = datetime.now(timezone.utc)
            data_ordenacao = datetime(2099, 1, 1)
            n_controle = item.get('numeroControlePNCP')
            cnpj = item.get('orgaoEntidade', {}).get('cnpj')
            cnpj_limpo = str(cnpj).replace('.', '').replace('/', '').replace('-', '')
            ano = item.get('anoCompra')
            sequencial = item.get('sequencialCompra')

            # Se tiver o número de controle, usamos o link v2 direto (Mais seguro)
            if n_controle:
                link_valido = f"https://pncp.gov.br/app/editais/v2/compra/{n_controle}"
            else:
                # Plano B: Caso o controle falhe, montamos a rota v2 manual
                link_valido = f"https://pncp.gov.br/app/editais/v2/compra/{cnpj_limpo}/{ano}/{sequencial}"
            if data_fim_str:
                try:
                    data_fim_utc = datetime.fromisoformat(data_fim_str).replace(tzinfo=timezone.utc)

                    if data_fim_utc<=agora:
                        continue
                    data_brasilia = data_fim_utc.astimezone(timezone(timedelta(hours=-3)))
                    data_formatada = data_brasilia.strftime('%d/%m/%Y %H:%M')
                    data_ordenacao=data_fim_utc
                except:
                    data_formatada = 'NA'
            else:
                data_formatada='NA'

            registro = {
                'DATA': data_formatada,
                'DATA_DT':data_ordenacao,
                'NUMERO': f"{item.get('numeroCompra')}/{item.get('anoCompra')}",
                'MODALIDADE': item.get('modalidadeNome', '').upper(),
                'ORGAO': item.get('orgaoEntidade', {}).get('razaoSocial', '').upper(),
                'OBJETO': objeto_limpo,
                'UASG': item.get('unidadeOrgao', {}).get('codigoUnidade', ""),
                'LINK':item.get('linkSistemaOrigem')
            }
            lista_temp.append(registro)
        lista_temp.sort(key=lambda x: x['DATA_DT'])
        for r in lista_temp:
            del r['DATA_DT']
            self.dados_filtrados.append(r)


    def gerar_planilha(self):
        caminho_arquivo=super().gerar_planilha()
        if not caminho_arquivo:
            return None
        try:

            wb = load_workbook(caminho_arquivo)
            ws = wb.active
            cores_modalidades = {
                'TOMADA DE PREÇOS': PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
                # Vermelho
                'PREGÃO - ELETRÔNICO': PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid"),
                # Verde
                'PREGÃO - PRESENCIAL': PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
                # Amarelo
                'CONCORRÊNCIA': PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid"),  # Laranja
                'CREDENCIAMENTO': PatternFill(start_color="800080", end_color="800080", fill_type="solid")  # Roxo
            }
            fonte_padrao = Font(name='Century Gothic', size=8)
            alinhamento = Alignment(wrap_text=True, vertical='center', horizontal='left')

            preenchimento_padrao = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                modalidade_celula = row[2]
                texto_modalidade = str(modalidade_celula.value).upper() if modalidade_celula.value else ""
                cor_localizada = False
                for cell in row:
                    cell.font = fonte_padrao
                    cell.alignment = alinhamento
                for chave, preenchimento in cores_modalidades.items():
                    if chave in texto_modalidade:
                        modalidade_celula.fill = preenchimento
                        cor_localizada = True
                        if chave in ['CREDENCIAMENTO', 'TOMADA DE PREÇOS']:
                            modalidade_celula.font = Font(name='Century Gothic', size=8, color="FFFFFF")
                        break
                if not cor_localizada:
                    modalidade_celula.fill = preenchimento_padrao

            ws.column_dimensions['A'].width = 18  # Data
            ws.column_dimensions['B'].width = 15  # Número
            ws.column_dimensions['C'].width = 25  # Modalidade
            ws.column_dimensions['D'].width = 40  # Órgão
            ws.column_dimensions['E'].width = 60  # Objeto
            ws.column_dimensions['F'].width = 12  # UASG
            ws.column_dimensions['G'].width = 24  # link

            wb.save(caminho_arquivo)


            return caminho_arquivo


        except Exception as e:
            print(f"Erro na estilização em bits: {e}")
            return caminho_arquivo




