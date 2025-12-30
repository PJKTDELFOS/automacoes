from email.mime.text import MIMEText
import requests
import pandas as pd
from datetime import datetime,timedelta
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import re
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import os
from keys import KEYS



def envio_de_email(arquivo_anexo):
    meu_email='boletinlicitacao@gmail.com'
    minha_senha=KEYS.senha
    email_destinho='albert.franca1992@gmail.com'

    msg=MIMEMultipart()
    msg['From']=meu_email
    msg['To']=email_destinho
    msg['Subject']=f'Cronograma diario PNCP-{datetime.now().strftime("%d/%m/%Y")}'
    corpo='Ola,\nSegue em anexo o cronograma de licitaçõe extraido do PNCP'
    msg.attach(MIMEBase('text', 'plain'))
    msg.attach(MIMEText(corpo,'plain'))

    with open (arquivo_anexo, 'rb') as arquivo:
        part=MIMEBase('application','octet-stream')
        part.set_payload(arquivo.read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', f"attachment; filename= {arquivo_anexo}")
        msg.attach(part)
    try:
        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.starttls()
        server.login(meu_email, minha_senha)
        server.sendmail(meu_email, email_destinho, msg.as_string())
        server.quit()
        print(f'Email enviado com sucesso! para {email_destinho}')
    except Exception as e:
        print(f"Erro ao enviar e-mail: {e}")



def cronograma():
    hoje = datetime.now()
    data_limite=hoje+ timedelta(days=15)
    data_final_para_api=data_limite.strftime("%Y%m%d")
    data_nome_arquivo = hoje.strftime('%d_%m_%Y')
    nome_arquivo_final = f'cronograma_{data_nome_arquivo}'
    pasta_projeto = r'C:\Users\Pichau\Desktop\PROJETOS\bot_pncp'
    nome_completo_arquivo =os.path.join(pasta_projeto,f'{nome_arquivo_final}.xlsx')
    base_url = "https://pncp.gov.br/api/consulta"
    endpoint = f"{base_url}/v1/contratacoes/proposta"
    palavraschave = ['site','pagina web','web','portal','software','institucional',]

    lista_final_filtrada=[]
    headers={
        'accept': 'application/json',
        "User-Agent": "Mozilla/5.0"
    }
    pagina_atual=1
    total_paginas=1


    try:
        while pagina_atual <= total_paginas:

            params={
                'dataFinal':data_final_para_api,
                 'uf':'RJ',  # paga pesquisas a nivel nacional so apagar o # do lado da uf
                'pagina':pagina_atual,
                'tamanhoPagina':50,#ajuste do tamanho da pagina, so vai ate 50
            }
            print(f'consultando pagina: {pagina_atual}')
            response=requests.get(endpoint,params=params,headers=headers)
            response.raise_for_status()
            dados_json=response.json()
            dados_brutos=dados_json.get('data',[])
            total_paginas=dados_json.get('totalPaginas',1)
            for item in dados_brutos:
                objeto=item.get('objetoCompra','')
                objeto_limpo=re.sub(r'[\x00-\x1F\x7F]', '', objeto)
                if any(palavra in objeto_limpo.lower() for palavra in palavraschave):
                    data_fim_str = item.get('dataEncerramentoProposta')
                    if data_fim_str:
                        data_fim=datetime.fromisoformat(data_fim_str.replace('Z','').split('.')[0])
                        if data_fim>hoje:
                            unidade = item.get('unidadeOrgao', {})or {}
                            registro = {
                                'DATA': data_fim.strftime('%d/%m/%Y %H:%M'),  # [cite: 507]
                                'NUMERO': f"{item.get('numeroCompra')}/{item.get('anoCompra')}",  # [cite: 504]
                                'MODALIDADE': item.get('modalidadeNome', ''),  # [cite: 507]
                                'ORGAO': item.get('orgaoEntidade', {}).get('razaoSocial', ''),  # [cite: 509]
                                'OBJETO': objeto_limpo.upper(),  # [cite: 507]
                                'UASG': unidade.get('codigoUnidade', "")  # [cite: 509]
                            }
                            lista_final_filtrada.append(registro)


            pagina_atual += 1
        if not lista_final_filtrada:
            print('nenhuma licitação encontrada com essas palavras chave')
            return
        df = pd.DataFrame(lista_final_filtrada)
        df['DATA_OBJ']=pd.to_datetime(df['DATA'],format='%d/%m/%Y %H:%M')
        df=df.sort_values(by='DATA_OBJ',ascending=True)
        df=df.drop(columns=['DATA_OBJ'])
        df = df[['DATA', 'NUMERO', 'MODALIDADE', 'ORGAO', 'OBJETO', 'UASG']]
        df.to_excel(nome_completo_arquivo, index=False)
        wb=load_workbook(nome_completo_arquivo)
        ws=wb.active
        cores_modalidades={
            'TOMADA DE PREÇOS': PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"), # Vermelho
            'PREGÃO - ELETRÔNICO': PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid"), # Verde
            'PREGÃO - PRESENCIAL': PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"), # Amarelo
            'CONCORRÊNCIA': PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid"), # Laranja
            'CREDENCIAMENTO': PatternFill(start_color="800080", end_color="800080", fill_type="solid") # Roxo
        }
        fonte_padrao=Font(name='Century Gothic',size=8)
        alinhamento=Alignment(wrap_text=True, vertical='center',horizontal='left')

        preenchimento_padrao = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            modalidade_celula=row[2]
            texto_modalidade=str(modalidade_celula.value).upper() if modalidade_celula.value else ""
            cor_localizada=False
            for cell in row:
                cell.font=fonte_padrao
                cell.alignment=alinhamento
            for chave, preenchimento in cores_modalidades.items():
                if chave in texto_modalidade:
                    modalidade_celula.fill=preenchimento
                    cor_localizada=True
                    if chave in ['CREDENCIAMENTO', 'TOMADA DE PREÇOS']:
                        modalidade_celula.font = Font(name='Century Gothic', size=8, color="FFFFFF")
                    break
            if not cor_localizada:
                modalidade_celula.fill = preenchimento_padrao




        ws.column_dimensions['A'].width = 18  # Data
        ws.column_dimensions['B'].width = 15  # Número
        ws.column_dimensions['C'].width = 25  # Modalidade
        ws.column_dimensions['D'].width = 40  # Órgão
        ws.column_dimensions['E'].width = 60  # Objeto
        ws.column_dimensions['F'].width = 12  # UASG
        wb.save(nome_completo_arquivo)



        print(f'\nSucesso! {len(df)} licitações filtradas encontradas.')
        print(f"Sucesso! Cronograma formatado em {nome_completo_arquivo}")
        envio_de_email(nome_completo_arquivo)
    except Exception as e:
        print(f"Erro ao processar dados da API: {e}")



if __name__ == "__main__":
    cronograma()